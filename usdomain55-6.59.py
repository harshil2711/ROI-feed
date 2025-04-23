import re
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

# === Setup Chrome ===
chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

# === Extract price from PDP ===
def extract_price(url):
    try:
        driver.get(url)
        time.sleep(10)
        price_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//div[@class='product-info-price']//span[@data-price-type='finalPrice']/span"))
        )
        price_text = price_element.text
        price = re.sub(r'[^\d.]', '', price_text)
        return f"{float(price):.2f}"  # Always return 2 decimal places like 24.00
    except Exception as e:
        print(f"❌ Failed to extract price for {url}: {e}")
        return 'N/A'

# === Load the Excel file ===
file_path = 'RESULT US.xlsx'
wb = load_workbook(file_path)
ws = wb['5500-6590']  # 👉 READ AND WRITE FROM Sheet2

# === Get headers and their column indices ===
headers = [cell.value for cell in ws[1]]
try:
    link_col = headers.index("Link") + 1
    expected_price_col = headers.index("Sale Price") + 1
except ValueError:
    raise Exception("Missing 'Link' or 'Sale Price' column in the Excel file.")

# === Add columns if needed ===
# Get accurate headers across all columns
# --- Accurate way to read headers ---
# --- Build header mapping (column name → column index) ---
headers_map = {}
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=1, column=col).value
    if header:
        headers_map[header.strip()] = col

# --- Get base columns ---
try:
    link_col = headers_map["Link"]
    expected_price_col = headers_map["Sale Price"]
except KeyError as e:
    raise Exception(f"Missing expected header: {e}")

# --- Insert 'PDP Price' right after 'Sale Price' ---
pdp_price_col = expected_price_col + 1
if "PDP Price" not in headers_map:
    ws.insert_cols(pdp_price_col)
    ws.cell(row=1, column=pdp_price_col).value = "PDP Price"
else:
    pdp_price_col = headers_map["PDP Price"]

# --- Insert 'Result' right after 'PDP Price' ---
result_col = pdp_price_col + 1
if "Result" not in headers_map:
    ws.insert_cols(result_col)
    ws.cell(row=1, column=result_col).value = "Result"
else:
    result_col = headers_map["Result"]


# === Process each row ===
from openpyxl.styles import numbers

# Inside your loop
for row_idx in range(2, ws.max_row + 1):
    pdp_value = ws.cell(row=row_idx, column=pdp_price_col).value
    if pdp_value not in [None, '', 'N/A']:
        continue  # Skip already processed rows

    url = ws.cell(row=row_idx, column=link_col).value
    print(f"➡️ Navigating to: {url}")
    expected_price_cell = ws.cell(row=row_idx, column=expected_price_col)
    expected_price = expected_price_cell.value

    price = extract_price(url)

    pdp_cell = ws.cell(row=row_idx, column=pdp_price_col)
    if price != 'N/A':
        pdp_cell.value = round(float(price), 2)
        pdp_cell.number_format = numbers.FORMAT_NUMBER_00
    else:
        pdp_cell.value = price

    if price != 'N/A' and expected_price is not None:
        try:
            result = 'Pass' if round(float(expected_price), 2) == round(float(price), 2) else 'Fail'
        except:
            result = 'Fail'
    else:
        result = 'Fail'

    ws.cell(row=row_idx, column=result_col).value = result
    wb.save("RESULT US.xlsx")

print("✅ Excel Sheet updated successfully!")
# === Close browser ===
driver.quit()


